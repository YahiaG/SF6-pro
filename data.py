import openpyxl, sqlite3, os
import tkinter as tk
from tkinter.filedialog import askopenfilename

def update_db():
    try:
        tk.Tk().withdraw() # part of the import if you are not using other tkinter functions
        path = askopenfilename(title='Choose SF6 File')
        print("You chose", path)
        print('.....')
        # Load Excel workbook
        wb_ob = openpyxl.load_workbook(path)

        # Select Active sheet
        sheet = wb_ob.active

        # Get number of columns & rows
        cols = sheet.max_column
        rows = sheet.max_row

        header = []
        for i in range(1, cols + 1):
            if sheet.cell(1,i).value:
                cell_value = ''.join(l for l in sheet.cell(1,i).value if l.isalnum())
            else:
                cell_value = f"header_{i}"
            header.append(cell_value)
        
        # Delete Database if exists
        try:
            os.remove("Links_DB.db")
        except:
            pass

        # Create database
        db = sqlite3.connect('Links_DB.db')
        cr = db.cursor()
        command = f"CREATE TABLE IF NOT EXISTS links({', '.join(header)})"
        cr.execute(command)
        for i in range(2, rows+1):
            rec = [str(sheet.cell(i, j).value).replace("'","") for j in range(1, cols+1)]
            insert_comm = "INSERT INTO links VALUES('" + "', '".join(rec) + "')"
            cr.execute(insert_comm)
        print("Database Updated")
        db.commit()
        db.close()
    except:
        print("Operation Failed")

#============================================================================

def search_links(site_id):
    db = sqlite3.connect('Links_DB.db')
    cr = db.cursor()

    head = ['SiteNoBEnd', 'SiteNoAEnd', 'TransmissionType', 'PathLengthKm', 'AntennaSizeBEnd', 'BAND', 'TxPowerdbm', 'ReceivedPowerBEndmdb', "Protection", 'Activity']
    for ind, text in enumerate(head):
        word = 30 if ind == 2 else 20
        print('\033[45m' + text.center(word) + '\033[0m', end=" ")
    print()
    links_list = []
    cr.execute(f"SELECT SiteNoBEnd, SiteNoAEnd, TransmissionType, PathLengthKm, AntennaSizeBEnd, BAND, TxPowerdbm, ReceivedPowerBEndmdb, Conc, Activity FROM links WHERE SiteNoBEnd = '{site_id}' ")
    result = cr.fetchall()
    for linkB in result:
        links_list.append(linkB)
    cr.execute(f"SELECT SiteNoBEnd, SiteNoAEnd, TransmissionType, PathLengthKm, AntennaSizeBEnd, BAND, TxPowerdbm, ReceivedPowerBEndmdb, Conc, Activity FROM links WHERE SiteNoAEnd = '{site_id}' ")
    result = cr.fetchall()
    for linkB in result:
        links_list.append(linkB)

    for lnk in links_list:
        if lnk[9] in ['Active', 'Planned', 'Accepted', 'Ready']:
            if lnk[9] != 'Active':
                print('\033[31m', end="")
            for ind, text in enumerate(lnk):
                word = 30 if ind == 2 else 20
                print(text.center(word), end=" ")
            print('\033[0m')

    db.commit()

    db.close()

#============================================================================

def affected_sites(site, route=[]):
    db = sqlite3.connect('Links_DB.db')
    cr = db.cursor()
    aff = [site]
    new_route = route + aff
    print(' >> '.join(new_route))
    cr.execute(f"SELECT SiteNoBEnd, Activity FROM links WHERE SiteNoAEnd = '{site}'")
    cascades = cr.fetchall()
    for x in cascades:
        if x[1] in ['Active', 'Planned', 'Accepted', 'Ready']:
            if x[0] in new_route:
                continue
            aff += affected_sites(x[0], new_route)
            
    db.close()
    return aff